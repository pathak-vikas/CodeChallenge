from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
import openpyxl

from restaurant.models import  Restaurant, Item

class Command(BaseCommand):
    help  = 'add sample data'

    def completed(self):
        self.stdout.write(self.style.SUCCESS('\nCommand Completed'))

    def handle(self, *args, **options):
        self.stdout.write(self.style.SUCCESS('Boom Ready for updates'))
        # import module
        
        # load excel with its path
        wrkbk = openpyxl.load_workbook("C:/Users/Evans Meja/Desktop/Restaurant/backend/web/web/apps/restaurant/management/commands/dummy/sample.xlsx")
        sh = wrkbk.active
        # iterate through excel and display data
        for i in range(2, sh.max_row+1):
            restaurant = sh.cell(row=i, column=2).value
            item = sh.cell(row=i, column=3).value
            price = sh.cell(row=i, column=4).value

            obj, created = Restaurant.objects.get_or_create(
                name=restaurant
            )
           
            item_obj, created = Item.objects.get_or_create(
                name=item,
                price=price,
                restaurant=obj
            )
        self.stdout.write(self.style.SUCCESS('Done'))